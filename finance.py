# -*- coding: Big5 -*-
###
#這段程式碼在計算完移動平均線和RSI指標後，使用rolling函數計算了20日的標準差，並根據標準差計算出了布林通道的上限和下限。然後，程式碼使用loc函數找出當股價低於布林通道下限且RSI指標小於30時的日期，並將signal欄位設置為1。這表示在這些日期，股價已經跌破了布林通道下限，且RSI指標顯示股票已經超賣，這可能是一個買入點。您可以根據這些信號進行進一步的分析和決策。
#加入line群組的發送指令:
###

import numpy as np
import pandas as pd
import yfinance as yf
import matplotlib.pyplot as plt
import datetime
from datetime import date, timedelta
##新增:將趨勢圖傳送到Line群組
import requests
from io import BytesIO
from PIL import Image
import lineTool
import json
import time
import sys
import os
#import xlwt             # pip install xlwt
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage




#from requests_toolbelt.multipart.encoder import MultipartEncoder                #pip install requests_toolbelt
#pip install xlsxwriter

# 設定Line Notify服務的存取權杖:  【股市etf群組】
access_token = 'optTVwv7UKvM7QEXFFTZNi9YB4yZZ1WojsWL6hxO2iVw2lCS7fAA/Uf9ztVqJXDb4xnXZdwuBZHgFK+eXz/bE86B8Ge+YBtEt6mEduMjFf5m7Ljqc7LdoDW1ss5BfC/iLJPbS+1UFv8yvk4zTCka1QdB04t89/1O/w1cDnyilFU=' # 改token
#access_token = 'OPkJV3'

# 設定股票代號
#symbol = '00878.TW'
# 設定股票代號列表
#symbols = ['00878.TW', '0056.TW', '00713.TW', '00919.TW']
# 設定股票代號字典

#關注的ETF
#symbols = {'00878.TW': '國泰永續高股息', '00919.TW': '群益台灣精選高息', '00713.TW': '元大台灣高息低波', '0056.TW': '元大高股息','006208.TW': '富邦台灣50', '00690.TW': '兆豐藍籌30', '00850.TW': '元大臺灣ESG永續', '00701.TW': '國泰股利精選30'}
#關注的金雞母績優股
symbols = {'2356.TW': '英業達','2330.TW': '台積電', '2454.TW': '聯發科', '2357.TW': '華碩', '6505.TW': '台塑','1102.TW': '亞泥','2324.TW': '仁寶'}
#關注的金融股
#symbols = {'2890.TW': '永豐金','2886.TW': '兆豐金','2885.TW': '元大金'}


#symbols = {'00900.TW': '高股息', '00888.TW': '高息', '00692.TW': '高息'}
#symbols = {'00878.TW': '國泰永續高股息', '00919.TW': '群益台灣精選高息', '00713.TW': '元大台灣高息低波', '0056.TW': '元大高股息'}
#symbols = {'00878.TW': '國泰永續高股息', '00919.TW': '群益台灣精選高息'}
#symbols = {'2330.TW': '台積電', '2454.TW': '聯發科', '2357.TW': '華碩', '6505.TW': '台塑'}
#symbols = {'006208.TW': '富邦台灣50', '00690.TW': '兆豐藍籌30', '00850.TW': '元大臺灣ESG永續', '00701.TW': '國泰股利精選30'}
#symbols = {'2324.TW': '仁寶', '2357.TW': '華碩', '1101.TW': '台泥', '2356.TW': '英業達'}
#symbols = {'1102.TW': '亞泥', '2886.TW': '兆豐金', '2890.TW': '永豐金', '2885.TW': '元大金'}
#symbols = {'1102.TW': '亞泥', '2890.TW': '永豐金'}
#symbols = {'00919.TW': '群益台灣精選高息'}
#symbols = {'00713.TW': '元大台灣高息低波'}
#symbols = {'0056.TW': '元大高股息'}
#symbols = {'00900.TW': '富邦高股息'}
#symbols = {'2890.TW': '永豐金'}
#symbols = {'2356.TW': '英業達'}
#symbols = {'0050.TW': '元大台灣50'}
#symbols = {'00690.TW': '兆豐藍籌30'}

'''
#方法1: 使用台灣證券交易所的股票代號轉換API，將股票代號轉換為中文名稱
url = 'https://mis.twse.com.tw/stock/api/getStockInfo.jsp?ex_ch=' + symbol
try:
    response = requests.get(url)
    # 將response內容轉為utf-8
    response.encoding = 'utf-8'
    name = response.json()['msgArray'][0]['n']
except requests.exceptions.JSONDecodeError:
    name = ""

#方法2: 使用台灣證券交易所的股票代號查詢網頁，獲取股票中文名稱
url = 'http://mis.twse.com.tw/stock/api/getStockInfo.jsp?c=' + symbol
response = requests.get(url)
pattern = re.compile(symbol + r'\s+(.*?)\s+')
match = pattern.search(response.text)
if match:
    symbol_name = match.group(1)
else:
    symbol_name = symbol      
#這個錯誤通常是由於API返回的數據格式不正確或無法解析而引起的。您可以嘗試使用try-except語句捕獲JSONDecodeError異常，並在捕獲到異常時處理它。
'''

#抓出今日日期
tonow = datetime.datetime.now()
Y1 = tonow.year
M1 = tonow.month
D1 = tonow.day
todaystr = str(Y1)+"/"+str(M1)+"/"+str(D1)

#for symbol in symbols:  #####################################連續依次查詢
'''
# 建立Excel檔案
# 建立一個空的Excel檔案
#writer = pd.ExcelWriter('stock_data.xlsx', engine='xlsxwriter')
# 檢查Excel檔案是否存在
if os.path.isfile('stock_data.xlsx'):
    # 如果存在，就開啟這個檔案
    ccdf = pd.read_excel('stock_data.xlsx', sheet_name=symbols[symbol])
    startrow = ccdf.shape[0]
    writer = pd.ExcelWriter('stock_data.xlsx', engine='openpyxl')
    writer.book = openpyxl.load_workbook('stock_data.xlsx')
else:
    # 如果不存在，就新建一個檔案
    startrow = 0
    writer = pd.ExcelWriter('stock_data.xlsx', engine='xlsxwriter')
'''
###-----------------------------------------------以下開始依個股跑回圈
for symbol, name in symbols.items():

    # 檢查檔案是否存在
    if os.path.isfile('999.xlsx'):
        # 如果檔案存在，載入它
        wb = load_workbook('999.xlsx')
        # 編輯特定的資料表
        if name in wb.sheetnames:
            ws = wb[name]
            # 在這裡進行資料表的編輯
        else:
            # 如果特定的資料表不存在，則建立一個新的資料表
            ws = wb.create_sheet(name)
            # 在這裡進行新資料表的初始化
            listtitle=['代號','股名','日期','收盤價(元)','成交量','RSI','MFI','ADL','Williams[20]','布林上限(元)','布林下限(元)','月均線(元)','季均線(元)','入場股價[45](元)','操作建議']
            ws.append(listtitle)
    else:
        # 如果檔案不存在，建立一個新的xlsx檔案
        wb = Workbook()
        # 建立特定的資料表
        ws = wb.create_sheet(name)
        # 在這裡進行新資料表的初始化
        listtitle=['代號','股名','日期','收盤價(元)','成交量','RSI','MFI','ADL','Williams[20]','布林上限(元)','布林下限(元)','月均線(元)','季均線(元)','入場股價[45](元)','操作建議']
        ws.append(listtitle)

    todaydate = ">>"+todaystr+"金融趨勢>>"+"\n"+" ["+str(symbol)+"]-[ "+name+" ] 分析摘要:"


    # 下載股價資料-最近1年
    df = yf.download(symbol, period='1y')
    #print(df)
    #df欄位格式: [  Date       Open       High        Low      Close  Adj Close     Volume   ]
    #sys.exit(1)

    # 計算移動平均線
    df['MA20'] = df['Close'].rolling(window=20).mean()
    df['MA60'] = df['Close'].rolling(window=60).mean()

    # 從資料中提取最高價（high）和最低價（low） 
    high_values = df['High'].values 
    low_values = df['Low'].values

    # 計算william指數 
    #df['WILLIAMS'] = (high_values - low_values.shift()) / low_values.shift() 
    # 將 low_values 轉換為 Pandas 的 Series
    #low_series = pd.Series(low_values)
    # 使用 shift() 方法
    #df['WILLIAMS'] = (high_values - low_series.shift()) / low_series.shift()
    #使用單純當日的WILLIAM指數的計算公式為：(最高價 - 收盤價) / 最高價 - 最低價) * -100
    #df['WILLIAMS'] = ((df['High'] - df['Close']) / (df['High'] - df['Low'])) * -100
    #使用過去20天週期的計算方式:WILLIAM指數的計算公式為：(過去45天收盤最高價-當日收盤價) / 過去45天最高價 - 過去45天最低價) * -100
    window_size = 20
    high_values = df['High'].rolling(window_size).max()
    low_values = df['Low'].rolling(window_size).min()
    df['WILLIAMS'] = ((high_values - df['Close'][-1]) / (high_values - low_values)) * 100


    # 計算MFI資金流向指標--------------------------------------------------------
    #有個MFI指標和RSI指標很類似，但它多考量了成交量的因素，而不像RSI只考量價格因素。
    #資金流量指標、資金流向指標(Money Flow Index , MFI)，是技術分析中用來衡量買賣壓力的工具，它與RSI相似但包含了成交量，而不像RSI只包括價格。
    #MFI被認為有一定的領先意義，通常用來預測市場走勢，透過MFI觀察出的超買、超賣訊號，可以幫投資人識別潛在的逆轉。
    #MFI高於80或以上要注意反轉的趨勢、當價格在強勁的下跌趨勢下繼續下跌時，MFI值可能低於20，則代表超賣的狀況。
    #如果價格走勢和MFI的指示相反，就會出現背離訊號，而失敗波動的看漲看跌也用來預測買入及賣出的機會。
    #MFI資金流向指標的計算公式為：100 - (100 / (1 + money_flow_ratio))
    #close_values = df['Close'].values 
    #df['FMI'] = (high_values + low_values + close_values) / 3
    #使用累?/派??（Accumulation/Distribution Line，ADL）??算。ADL是一种基于成交量和价格?化的技?指?，用于衡量??的???力。MFI是基于ADL?算的，因此也被??MFI/ADL指?。
    # ?算ADL
    adl = ((df['Close'] - df['Low']) - (df['High'] - df['Close'])) / (df['High'] - df['Low']) * df['Volume']
    adl = adl.cumsum()

    #df['Volume'] = df['Volume'].fillna(0)
    # 計算 MFI
    typical_price = (df['High'] + df['Low'] + df['Close']) / 3
    money_flow = typical_price * df['Volume']
    positive_flow = np.where(typical_price > typical_price.shift(1), money_flow, 0)
    negative_flow = np.where(typical_price < typical_price.shift(1), money_flow, 0)
    positive_flow_sum = pd.Series(positive_flow).rolling(window=14).sum()
    negative_flow_sum = pd.Series(negative_flow).rolling(window=14).sum()
    money_flow_ratio = np.where(negative_flow_sum == 0, 0.5, positive_flow_sum / negative_flow_sum)
    mfi = 100 - (100 / (1 + money_flow_ratio))
    # 中
    df['MFI'] = mfi
    df['ADL'] = adl
    #檢查MFI指標是否計算成功
    print(df['MFI'])
    print(df['ADL'])
    #----------------------------------------------------------------------------------------
    #sys.exit(1)

    # 計算RSI指標
    delta = df['Close'].diff()
    gain = delta.where(delta > 0, 0)
    loss = -delta.where(delta < 0, 0)
    avg_gain = gain.rolling(window=14).mean()
    avg_loss = loss.rolling(window=14).mean()
    rs = avg_gain / avg_loss
    df['RSI'] = 100 - (100 / (1 + rs))
    #print(df['RSI'][-1])

    # 計算布林通道
    df['std'] = df['Close'].rolling(window=20).std()
    df['upper'] = df['MA20'] + 2 * df['std']
    df['lower'] = df['MA20'] - 2 * df['std']

    #手動設定買入價格和停損價格
    #buy_price = 17.5
    #stop_loss_price = 18.5

    '''
    # 利用最近一天的收盤價，自動計算最佳買入價格和停損價格
    df['diff'] = df['upper'] - df['lower']
    df['buy_price'] = df['lower'] + 0.2 * df['diff']
    df['stop_loss_price'] = df['lower'] + 0.1 * df['diff']
    #print('建議之最佳買入點:' + str(df['buy_price']))
    '''
    # 利用前一個月的布林通道上下限來計算最佳買入價格和停損價格
    df['diff'] = df['upper'] - df['lower']
    df['buy_price'] = df['lower'].rolling(window=45).mean() + 0.2 * df['diff'].rolling(window=45).mean()
    #這段程式碼是用來計算最佳買入價格的公式，其中 df['lower'].rolling(window=20).mean() 是指計算過去45天的布林通道下限平均值，而 
    #df['diff'].rolling(window=20).mean() 是指計算過去45天的布林通道上下限區間差的平均值。因此，這段程式碼的意思是：利用過去45天的布林通道下限平均值加上過去45天的布林通道上下限區間差的平均值乘上0.2，來計算最佳買入價格。
    #這個公式的目的是要讓投資人在股價處於低點時買進，並且設定一個停損價格，以控制風險。這個公式的核心是利用布林通道的下限來判斷股價是否處於低點，再加上一個緩衝區間，以避免買進時股價已經反彈。而這個緩衝區間的大小，就是由0.2倍的布林通道上下限區間差來決定的。
    df['stop_loss_price'] = df['lower'].rolling(window=45).mean() + 0.1 * df['diff'].rolling(window=45).mean()

    # 找出最佳買入點
    df['signal'] = 0
    #df.loc[(df['Close'] < df['lower']) & (df['RSI'] < 30) & (df['Close'] <= buy_price), 'signal'] = 1
    #df.loc[(df['Close'] < df['lower']) & (df['RSI'] < 30) & (df['Close'] <= df['buy_price']), 'signal'] = 1
    df.loc[(df['Close'] < df['lower'] * 1.05) & (df['RSI'] < 30) & (df['Close'] <= df['buy_price']), 'signal'] = 1
    #這段程式碼是用來判斷是否有買入訊號的條件，其中包括股價低於布林通道下軌線（lower）的1.05倍、RSI指標小於30、且股價小於等於買入價格（buy_price）。如果符合這些條件，則會在對應的資料列中標記signal為1，表示有買入訊號。
    #至於如何決定最佳買入點，這通常需要考慮多種因素，例如技術指標、基本面分析、市場趨勢等等。這些因素可以根據個人的投資策略和風險偏好進行綜合考慮，以找出最佳的買入點。建議在進行投資前，先進行充分的研究和分析，並制定出明確的投資策略和風險控制措施。

    '''
    # 判斷是否該買進、賣出或觀望
    last_close = df['Close'][-1]
    if df['signal'][-1] == 1:
        print(symbol+f'最近一天收盤價已低於布林下限且RSI指標小於30，建議買進，買入價格為{last_close:.2f}元')
        print(symbol+f'建議設定停損價格為{stop_loss_price:.2f}元')
    elif last_close > df['MA20'][-1] and last_close > df['upper'][-1]:
        print(symbol+f'最近一天收盤價已高於中線且接近布林上限，建議賣出，賣出價格為{last_close:.2f}元')
    else:
        print(symbol+'最近一天收盤價在布林通道內，建議觀望。')
    '''
    # 判斷是否該買進、賣出或觀望
    last_close = df['Close'][-1]
    #print('last_close:'+str(last_close))
    last_rsi = df['RSI'][-1]
    #print('last_rsi:'+str(last_rsi))
    last_buy_price = df['buy_price'][-1]
    print('利用最近45天交易數據推算的最佳買點:'+ str(last_buy_price))
    last_stop_loss_price = df['stop_loss_price'][-1]
    print('停損價格為:'+str(last_stop_loss_price))

    if df['signal'][-1] == 1:
        keyword = symbol+f'最近一天收盤價[{last_close:.2f}元]，已經出現買入訊號了!'+'\n'+'**買入訊號設定條件為: 股價已低於布林通道下軌線的1.05倍、且RSI指標小於30、且股價低於估算的合理買入價格。'+'\n'+'建議趕快買進，買入參考價格為'+str(round(last_buy_price*0.995,2))+'元。(祝發大財~)'+'\n'+'另提供參考停損價格為'+str(round(last_stop_loss_price,2))+'元。'    
    elif last_close >= df['MA20'][-1] and last_close >= df['MA60'][-1] and last_close >= df['upper'][-1]*0.995:
        keyword = symbol+f'最近一天收盤價[{last_close:.2f}元]，已高於MA20月均線及MA60季均線且超過布林上限。'+'\n'+'強烈建議可賣出。'+'\n'+'賣出參考價格:'+str(round(last_close*1.005,2))+'元。(發大財了~)'+'\n'+'RSI指標為'+str(round(last_rsi,2))   
    elif last_close >= df['MA20'][-1] and last_close >= df['upper'][-1]*0.995:
        keyword = symbol+f'最近一天收盤價[{last_close:.2f}元]，已高於MA20月均線且超過布林上限。'+'\n'+'若有賣出計畫，建議可賣出。'+'\n'+'賣出參考價格為'+str(round(last_close*1.005,2))+'元。'+'\n'+'RSI指標為'+str(round(last_rsi,2)) 
    elif last_close <= df['MA20'][-1] and last_close <= df['MA60'][-1] and last_close <= df['lower'][-1]*1.005:
        keyword = symbol+f'最近一天收盤價[{last_close:.2f}元]，已低於MA20月均線及MA60季均線且低於布林下限。'+'\n'+'強烈建議可加碼買進!!! (讓您發財~)'+'\n'+'RSI指標為'+str(round(last_rsi,2))    
    elif last_close <= df['MA20'][-1] and last_close <= df['lower'][-1]*1.005:
        keyword = symbol+f'最近一天收盤價[{last_close:.2f}元]，已低於MA20月均線且低於布林下限。'+'\n'+'若有要加碼買進的計畫，建議可以準備了喔。 (讓您發財~)'+'\n'+'RSI指標為'+str(round(last_rsi,2))    
    #elif last_close >= df['upper'][-1] * 0.95:
    #    keyword = symbol+f'最近一天收盤價[{last_close:.2f}元]，已高於布林上限0.95的緩衝區間。'+'\n'+'若有要賣出的計畫，建議可以準備了。'+'\n'+'RSI指標為'+str(round(last_rsi,2))    
    else:
        keyword = symbol+f'最近一天收盤價[{last_close:.2f}元]，還在布林通道內。'+'\n'+'建議持續觀望~'+'\n'+'RSI指標為'+str(round(last_rsi,2))
        

    #RS指標是一種技術分析指標，用於衡量股票的超買和超賣程度。RSI指標的計算方式是根據一段時間內股票漲跌幅度的平均值，來計算股票的強弱程度。RSI指標的取值範圍為0到100，一般認為RSI指標在30以下表示股票處於超賣狀態，而在70以上表示股票處於超買狀態。
    #在程式碼中建議RSI指標小於30時買進股票，是因為當RSI指標小於30時，表示股票處於超賣狀態，市場情緒可能已經過度悲觀，股票價格可能已經被低估。此時，如果股票的基本面沒有發生重大變化，可能是一個買入的好時機。當然，這只是一個參考，投資者還需要根據自己的風險承受能力和投資目標，做出更全面和準確的決策。
    '''
    ####準備將結果存入EXCEL中
    # 將每個股票代號的資料寫入Excel檔案
    # 讀取現有的資料表，如果不存在就建立一個新的
    '''
    #會建立一個名為stock_data.xlsx的Excel檔案，並逐一處理symbols字典中的每個股票代號。
    #對於每個股票，它會下載最近1年的股價資料，計算移動平均線、MFI和ADL，然後將資料寫入Excel檔案中。每個股票的資料表名稱會以股票名稱命名，欄位格式也會設定好。最後，它會儲存Excel檔案。

    
    # 將資料寫入Excel檔案
    #下面會寫入所有的資料
    #df.to_excel(writer, sheet_name=sheet_name, index=False)
    # 將所有資料寫入Excel檔案
    #df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)    

    # 新增欄位格式
    '''
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    date_format = workbook.add_format({'num_format': 'yyyy-mm-dd'})
    number_format = workbook.add_format({'num_format': '#,##0.00'})
    worksheet.set_column('A:A', 12, date_format)
    worksheet.set_column('B:B', None, number_format)
    worksheet.set_column('C:C', None, number_format)
    worksheet.set_column('D:D', None, number_format)
    worksheet.set_column('E:E', None, number_format)
    worksheet.set_column('F:F', None, number_format)
    worksheet.set_column('G:G', None, number_format)
    worksheet.set_column('H:H', None, number_format)
    # 取得最近一天的資料
    latest_data = df.iloc[-1:]

    #將最近一天的資料寫入Excel檔案
    latest_data.to_excel(writer, sheet_name=symbol, startrow=startrow+1, index=False, header=True)
    # 將資料寫入Excel檔案
    #df.to_excel(writer, sheet_name=symbol, startrow=startrow, index=False)
    '''
    # 在這裡進行excel其他的資料表編輯
    # 逐一檢查每個 C 欄位的值了。從第二列開始，從第三欄到第三欄（也就是 C 欄位）逐一取得每個儲存格
    c2_value = ""
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            if cell.value == todaystr:
                # 如果 C 欄位的值等於 'EEE'，執行這裡的程式碼塊
                c2_value = 'YES'
                break
            else:
                # 如果 C 欄位的值不等於 'EEE'，執行這裡的程式碼塊
                c2_value = 'NO'

    
    # 判斷 C2 儲存格的值是否等於 'todaystr'   表示今天的資料已經寫過了
    '''
    if c2_value is None:
        # 如果不等於 todaystr，表示該日期的資料不存在，則繼續新增一列
        # 在第二列插入一行，永遠寫入第一列
        #ws.insert_rows(1)
        #listtitle=['代號','股名','日期','收盤價(元)','成交量','RSI','MFI','ADL','Williams[20]','布林上限(元)','布林下限(元)','月均線(元)','季均線(元)','入場股價[45](元)','操作建議']
        listdatas=[symbol , name , todaystr , str(round(last_close,2)) , str(round(df['Volume'][-1],2)) , str(round(last_rsi,2)) , str(round(df['MFI'][-1],2)) , str(round(df['ADL'][-1]/10000,2)) , str(round(df['WILLIAMS'][-1],2))+'%' , str(round(df['upper'][-1],2)) , str(round(df['lower'][-1],2)) , str(round(df['MA20'][-1],2)) , str(round(df['MA60'][-1],2)) , str(round(last_buy_price,2)) , keyword]
        # 將資料寫入第二列
        ws.append(listdatas)    
    '''
    if c2_value == 'YES':
        # 如果等於 todaystr，表示胎日期資料已存在，則用修改的方式覆寫資料
        # 在第二列插入一行，永遠寫入第一列
        # 修改第二列的資料
        ws.cell(row=2, column=1, value=symbol)
        ws.cell(row=2, column=2, value=name)
        ws.cell(row=2, column=3, value=todaystr)
        ws.cell(row=2, column=4, value=str(round(last_close,2)))
        ws.cell(row=2, column=5, value=str(round(df['Volume'][-1],2)))
        ws.cell(row=2, column=6, value=str(round(last_rsi,2)))
        ws.cell(row=2, column=7, value=str(round(df['MFI'][-1],2)))
        ws.cell(row=2, column=8, value=str(round(df['ADL'][-1]/10000,2)))
        ws.cell(row=2, column=9, value=str(round(df['WILLIAMS'][-1],2))+'%')
        ws.cell(row=2, column=10, value=str(round(df['upper'][-1],2)))
        ws.cell(row=2, column=11, value=str(round(df['lower'][-1],2)))
        ws.cell(row=2, column=12, value=str(round(df['MA20'][-1],2)))
        ws.cell(row=2, column=13, value=str(round(df['MA60'][-1],2)))
        ws.cell(row=2, column=14, value=str(round(last_buy_price,2)))
        ws.cell(row=2, column=15, value=keyword)
    
    else:
        # 如果不等於 todaystr，表示該日期的資料不存在，則繼續新增一列
        # 在第二列插入一行，永遠寫入第一列
        #ws.insert_rows(1)
        #listtitle=['代號','股名','日期','收盤價(元)','成交量','RSI','MFI','ADL','Williams[20]','布林上限(元)','布林下限(元)','月均線(元)','季均線(元)','入場股價[45](元)','操作建議']
        listdatas=[symbol , name , todaystr , str(round(last_close,2)) , str(round(df['Volume'][-1],2)) , str(round(last_rsi,2)) , str(round(df['MFI'][-1],2)) , str(round(df['ADL'][-1]/10000,2)) , str(round(df['WILLIAMS'][-1],2))+'%' , str(round(df['upper'][-1],2)) , str(round(df['lower'][-1],2)) , str(round(df['MA20'][-1],2)) , str(round(df['MA60'][-1],2)) , str(round(last_buy_price,2)) , keyword]
        # 將資料寫入第二列
        ws.append(listdatas)

    # 儲存檔案
    #wb.save('999.xlsx')


    ####丟出line夜首-----------------------------------------------------------------------------------
    msgsouce = "近45天交易數據的最佳買點: "+ str(round(last_buy_price,2)) + "元"+"\n"
    msgsouce = msgsouce + "*布林通道上限: " + str(round(df['upper'][-1],2)) + "\n"
    msgsouce = msgsouce + "*月均線參考值: " + str(round(df['MA20'][-1],2)) + "\n"
    msgsouce = msgsouce + "*季均線參考值: " + str(round(df['MA60'][-1],2)) + "\n"
    msgsouce = msgsouce + "*布林通道下限: " + str(round(df['lower'][-1],2)) + "\n"+ "\n"
    msgsouce = msgsouce + "*Williams[20]:    " + str(round(df['WILLIAMS'][-1],2)) +"%"+"\n"
    msgsouce = msgsouce + "*MFI資金流向指標:  " + str(round(df['MFI'][-1],2))+"\n"
    msgsouce = msgsouce + "*A/D Line指標:  " + str(round(df['ADL'][-1]/10000,2))    
    msg = "\n"+ todaydate +"\n"+ str(msgsouce)
    r00 = lineTool.lineNotify(access_token,msg)


    # 繪製趨勢圖
    fig, ax = plt.subplots(figsize=(16, 9))

    ax.plot(df.index, df['Close'], lw=3,marker='.',color='#FF0000', label='C_Price')
    ax.plot(df.index, df['MA20'], lw=3,color='#EE7700', linestyle='-', label='MA20 day')
    ax.plot(df.index, df['MA60'], lw=3,color='#7700BB', linestyle='-', label='MA60 day')
    ax.plot(df.index, df['upper'], color='k', linestyle='--', label='Bollinger Bands_Upper')
    ax.plot(df.index, df['lower'], color='k', linestyle='--', label='Bollinger Bands_Lower')

    ax.fill_between(df.index, df['upper'], df['lower'], alpha=0.2)
    #下面是標記最近一天'的收盤價:
    #ax.axhline(df['Close'][-1], color='b', linestyle='--', label='Last Close')
    #ax.axhline(df['upper'][-1], color='r', linestyle='--', label='Upper')
    #ax.axhline(df['lower'][-1], color='g', linestyle='--', label='Lower')

    # 顯示最近一天收盤價的RSI值
    ax2 = ax.twinx()
    ax2.plot(df.index, df['RSI'], lw=1,color='b', label='RSI')
    ax2.axhline(y=30, lw=1,color='g', linestyle='--')
    ax2.axhline(y=70, lw=1,color='g', linestyle='--')
    ax2.set_ylim([0, 100])
    ax2.set_ylabel('RSI')
    ax2.legend(loc='upper left')

    ax.legend()
    ax.set_title(symbol+'~'+todaystr)
    ax.set_xlabel('Date')
    ax.set_ylabel('Price')

    # 讀取趨勢圖轉換為將其保存到一個BytesIO對象中
    '''
    fig, ax9 = plt.subplots()
    ax9.plot(df.index, df['Close'])
    ax9.set_title('Stock Price')
    ax9.set_xlabel('Date')
    ax9.set_ylabel('Price')
    '''
    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    #image = Image.open(buffer)   

    #plt.show()


    ###
    #這段程式碼使用了matplotlib庫來繪製趨勢圖。程式碼會繪製股價、20日和60日的移動平均線，以及布林通道的上限和下限。程式碼還使用fill_between函數填充了布林通道的區域。最後，程式碼顯示了繪製好的趨勢圖。
    ###
    '''
    # 蒐集最近一天的收盤價
    last_close = df['Close'][-1]
    # 判斷是否該買進、賣出或觀望
    if last_close < df['MA20'][-1] and last_close < df['lower'][-1]:
        print(symbol+'最近一天收盤價已低於中線且接近布林下限，建議買進。')
    elif last_close > df['MA20'][-1] and last_close > df['upper'][-1]:
        print(symbol+'最近一天收盤價已高於中線且接近布林上限，建議賣出。')
    else:
        print(symbol+'最近一天收盤價在布林通道內，建議觀望。')
    '''



    # 將趨勢圖轉換為二進制數據
    image_binary = buffer.getvalue()

    # 傳送趨勢圖到Line群組
    url = 'https://notify-api.line.me/api/notify'
    headers = {'Authorization': 'Bearer ' + access_token}
            #data = {'message': 'Stock Price', 'imageFile': ('image.png', buffer, 'image/png')}
            #response = requests.post(url, headers=headers, files=data)
    data = {'message': "\n"+keyword}
    files = {'imageFile': ('image.png', image_binary, 'image/png')}
    response = requests.post(url, headers=headers, data=data, files=files)     #最正確語法

    # 檢查HTTP響應
    print(response.status_code)
    print(response.text)

    #我們首先讀取趨勢圖並將其保存到一個BytesIO對象中。然後，我們使用BytesIO對象的getvalue()
    #方法將趨勢圖轉換為二進制數據。接下來，我們使用HTTP請求向Line Notify API發送消息，包括消息標題和趨勢圖。在HTTP請求中，我們將趨勢圖的二進制數據作為文件傳送，並將消息標題作為表單數據傳送。最後，我們可以檢查HTTP響應以確定消息是否成功發送。

    #在上面的程式碼中，我們首先設定Line Notify服務的存取權杖，然後讀取趨勢圖並將其轉換為PNG格式。接下來，我們使用HTTP請求向Line發送通知消息，包括消息標題和趨勢圖。最後，我們可以檢查HTTP響應以確定消息是否成功發送。
    #請注意，Line Notify服務有一些限制，例如每天最多只能發送1000條消息，每個消息最大只能包含1MB的數據等。如果您需要發送更多的消息或更大的數據，請考慮使用其他類似的服務，例如Telegram Bot等。
    msg_end = todaystr+"///////發財"
    r99 = lineTool.lineNotify(access_token,msg_end)

    #將圖檔寫入EXCEL中
    '''
    # 先刪除工作表中原先已經存在的圖片，避免檔案過大
    for imga in ws._images:
        if imga.anchor == 'A15':
            del ws._images[imga]
    '''
    # 刪除工作表中的所有圖片
    ws._images = []
    
    # 將圖片寫入工作表中
    imgg = XLImage(BytesIO(image_binary))
    ws.add_image(imgg, 'A15')
    # 儲存EXCEL檔案
    wb.save('999.xlsx')

    time.sleep(5)  # 暫停5秒鐘




